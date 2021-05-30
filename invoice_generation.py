import openpyxl as xl
from datetime import date
from num2words import num2words

count = 1
row = 12
i=0
total_val = 0

def add_space():
	for i in range(3):
		print()

def check_empty(x):
	is_empty = bool(x)
	return is_empty


add_space()
	
today_date = date.today()
print(today_date)

add_space()

wb = xl.load_workbook('Invoice Format.xlsx')
sheet = wb['Sheet 1']


client_name = sheet.cell(7,2)
client_address = sheet.cell(8,2)
invoice_no = sheet.cell(7,6)
dt = sheet.cell(8,6)


client_name.value = input('Enter name of client -> ')
client_address.value = input("Enter client's address -> ")
invoice_no.value = input('Invoice no. -> ')
dt.value = input("Date (Leave empty for today's date) -> ")

if not check_empty(dt.value):
	dt.value = today_date	

total_items = int(input('Enter total number of entries -> '))

add_space()

while count<=total_items:
	
	s_no = sheet.cell(row,1)
	item_name = sheet.cell(row,2)
	quantity = sheet.cell(row,4)
	price = sheet.cell(row,5)
	amount = sheet.cell(row,6)	
	s_no.value = count
	item_name.value = input('Enter item description -> ')
	quantity.value = input('Quantity -> ')
	price.value = input('Price -> ')
	amount.value = int(quantity.value)*int(price.value)
	count+=1
	row+=1
	add_space()
	total_val = total_val + amount.value

gst = int(input('Enter GST % -> '))


total = sheet.cell(27,6)
cgst = sheet.cell(28,6)
sgst = sheet.cell(29,6)
final = sheet.cell(30,6)
amount_in_words = sheet.cell(30,3)

total.value = total_val
cgst.value = gst*total_val/200
sgst.value = gst*total_val/200
final.value = total.value + cgst.value + sgst.value
num = int(final.value)

amount_in_words.value = num2words(final.value)

wb.save('Invoice No ' + invoice_no.value + '.xlsx')
